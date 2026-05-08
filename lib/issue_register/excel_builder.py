# -*- coding: utf-8 -*-
"""
Build the issue register Excel workbook by filling data into the LB template.

The template (lib/template.xltx) already contains all formatting for:
  - Rows 1-11  (header, key, disclaimer, distribution block, column headers)
  - Rows 12+   (pre-formatted data rows)
  - Date columns L+ (pre-formatted empty cells)

We load the template, clear/overwrite content, and save.
"""

import os
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.styles.colors import Color
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import XDRPositiveSize2D

# Header dark fill (theme 0 / dark1 = black, tint -0.15 → near-black matching LB template)
_HEADER_FILL = PatternFill(patternType='solid',
                           fgColor=Color(theme=0, tint=-0.14999847407452621))

_TEMPLATE = os.path.join(os.path.dirname(__file__), 'template.xltx')
_LOGO     = os.path.join(os.path.dirname(__file__), 'lb_logo.png')

# ── Layout constants (describe the template; effective values are computed at runtime) ──
HEADER_ROW      = 11  # row of column-label headers in the template
DATA_ROW_START  = 12  # first data row in the template
FIRST_DATE_COL  = 12  # column L
_DIST_FIRST_ROW = 4   # first distribution (recipient) row
_DIST_TMPL_ROWS = 7   # number of distribution rows pre-formatted in the template
_HEADER_PADDING = 4   # blank rows inserted between the header row and the first data row
                      # so that column-label cells can be merged tall enough for rotated text

_HEADER_LABELS = [
    'Drawing Package', 'Project', 'Originator',
    'Functional Breakdown', 'Spatial Breakdown', 'Form', 'Discipline',
    'Number', 'Document Title', 'Size', 'Scale',
]

_DATA_ROW_HEIGHT = 21.6


# ── Date helpers ──────────────────────────────────────────────────────────────

def _parse_date(date_str):
    for fmt in ('%d/%m/%y', '%d/%m/%Y', '%d.%m.%y', '%d.%m.%Y'):
        try:
            return datetime.strptime(date_str, fmt)
        except ValueError:
            pass
    return None


def _fmt_header(date_str):
    dt = _parse_date(date_str)
    return dt.strftime('%d/%m/%y') if dt else date_str


def _fmt_title(date_str):
    dt = _parse_date(date_str)
    return dt.strftime('%d.%m.%Y') if dt else date_str


# ── Style cloning helpers ─────────────────────────────────────────────────────

def _copy_cell_style(src, dst):
    """Copy font, fill, border, alignment from src cell to dst cell."""
    if src.has_style:
        dst.font      = src.font.copy()
        dst.fill      = src.fill.copy()
        dst.border    = src.border.copy()
        dst.alignment = src.alignment.copy()


def _snapshot_style(cell):
    """Capture all style properties as independent copies before the cell is modified."""
    if not cell.has_style:
        return None
    return {
        'font':          cell.font.copy(),
        'fill':          cell.fill.copy(),
        'border':        cell.border.copy(),
        'alignment':     cell.alignment.copy(),
        'number_format': cell.number_format,
    }


def _apply_snapshot(cell, snap):
    """Apply a style snapshot captured by _snapshot_style."""
    if snap is None:
        return
    cell.font          = snap['font']
    cell.fill          = snap['fill']
    cell.border        = snap['border']
    cell.alignment     = snap['alignment']
    cell.number_format = snap['number_format']


def _unmerge_region(ws, min_row, max_row, min_col, max_col):
    """Remove all merges overlapping a region and flush stale MergedCell cache entries.

    openpyxl leaves MergedCell (read-only) objects in ws._cells after removing
    a merge range. Deleting them forces fresh writable Cell objects on next access.
    """
    to_remove = [
        m for m in list(ws.merged_cells.ranges)
        if m.min_row <= max_row and m.max_row >= min_row
        and m.min_col <= max_col and m.max_col >= min_col
    ]
    for m in to_remove:
        ws.merged_cells.remove(m)
        for r in range(m.min_row, m.max_row + 1):
            for c in range(m.min_col, m.max_col + 1):
                ws._cells.pop((r, c), None)




# ── Main entry point ──────────────────────────────────────────────────────────

def build_register(sheets_data, issue_keys, settings, output_path, project_info):
    wb = load_workbook(_TEMPLATE)
    ws = wb.active

    n_dates  = len(issue_keys)
    last_col = FIRST_DATE_COL + n_dates - 1

    # ── Compute effective row layout ──────────────────────────────────────
    # Distribution block: rows 4 .. (eff_header_row - 1)
    # Header row:         eff_header_row
    # Padding rows:       eff_header_row+1 .. eff_data_start-1  (merged with header)
    # Data rows:          eff_data_start ..
    n_recipients    = len(settings.get('recipients', []))
    extra_dist_rows = max(0, n_recipients - _DIST_TMPL_ROWS)
    eff_header_row  = HEADER_ROW + extra_dist_rows
    eff_data_start  = eff_header_row + 1 + _HEADER_PADDING

    # Snapshot styles from TEMPLATE rows BEFORE any modification/insertion.
    _date_data_snap = _snapshot_style(ws.cell(row=DATA_ROW_START, column=FIRST_DATE_COL))
    _date_hdr_snap  = _snapshot_style(ws.cell(row=HEADER_ROW,     column=FIRST_DATE_COL))
    _date_r3_snap   = _snapshot_style(ws.cell(row=3,              column=FIRST_DATE_COL))
    _date_col_width = (9.0 if settings.get('suitability_enabled')
                       else ws.column_dimensions[get_column_letter(FIRST_DATE_COL)].width or 4.57)

    # How many date columns are already in the template?
    template_date_cols = ws.max_column - (FIRST_DATE_COL - 1)

    # ── Insert extra distribution rows before the header row ─────────────
    if extra_dist_rows > 0:
        ws.insert_rows(HEADER_ROW, extra_dist_rows)
        last_tmpl_dist    = HEADER_ROW - 1  # row 10 — unchanged after insert_rows above
        # Use the second-to-last template dist row (row 9) as the style reference
        # so inserted rows inherit full internal gridlines rather than any clean-edge
        # styling that may be baked into the template's last distribution row.
        style_tmpl_dist   = max(_DIST_FIRST_ROW, last_tmpl_dist - 1)
        for new_r in range(HEADER_ROW, eff_header_row):
            ws.row_dimensions[new_r].height = (
                ws.row_dimensions[last_tmpl_dist].height or 15)
            for c in range(1, FIRST_DATE_COL + template_date_cols):
                _copy_cell_style(ws.cell(row=style_tmpl_dist, column=c),
                                 ws.cell(row=new_r, column=c))
                # Template cols A-H may be merged cells (no style); force grey fill.
                ws.cell(row=new_r, column=c).fill = _HEADER_FILL

    # ── Insert padding rows between header and first data row ─────────────
    # Each column-label cell is later merged from eff_header_row down through
    # all padding rows, giving the combined height needed for rotated text.
    ws.insert_rows(eff_header_row + 1, _HEADER_PADDING)
    for pad_r in range(eff_header_row + 1, eff_data_start):
        ws.row_dimensions[pad_r].height = _DATA_ROW_HEIGHT
        for c in range(1, FIRST_DATE_COL + template_date_cols):
            _copy_cell_style(ws.cell(row=eff_header_row, column=c),
                             ws.cell(row=pad_r, column=c))

    # ── Expand date columns if we need more than the template provides ────
    if n_dates > template_date_cols:
        _expand_date_columns(ws, template_date_cols, n_dates,
                             eff_header_row, eff_data_start)

    # ── Set ALL date column widths to a consistent value ─────────────────
    for _ci in range(n_dates):
        ws.column_dimensions[get_column_letter(FIRST_DATE_COL + _ci)].width = _date_col_width

    # ── Widen columns B-H ────────────────────────────────────────────────
    for _col, _w in [('B', 9), ('C', 9), ('D', 9), ('E', 9), ('F', 9), ('G', 9), ('H', 22)]:
        ws.column_dimensions[_col].width = _w


    # ── Row 3: label merged I:K (cols 9-11), right-aligned flush against date cols ──
    _r3_lbl_snap = _snapshot_style(ws.cell(row=3, column=9))
    _unmerge_region(ws, 3, 3, 9, last_col)
    ws.merge_cells(start_row=3, start_column=9, end_row=3, end_column=11)
    _lbl = ws.cell(row=3, column=9)
    if _r3_lbl_snap:
        _apply_snapshot(_lbl, _r3_lbl_snap)
        _f = _r3_lbl_snap['font']
        _lbl.font = Font(name=_f.name, size=_f.size, bold=True, italic=_f.italic,
                         color=_f.color, underline=_f.underline, strike=_f.strikethrough)
    _lbl.value     = 'Issue date & revision'
    _lbl.alignment = Alignment(horizontal='right', vertical='center', wrap_text=False)

    # Row 3 date cells: first column shows register issue date+revision from settings.
    _reg_date = settings.get('register_issue_date', '').strip()
    _reg_rev  = settings.get('register_revision', '').strip()
    for _ci, (_ds, _) in enumerate(issue_keys):
        _c = ws.cell(row=3, column=FIRST_DATE_COL + _ci)
        _apply_snapshot(_c, _date_r3_snap)
        if _ci == 0:
            _val = ' | '.join(x for x in [_reg_date, _reg_rev] if x)
            _c.value     = _val if _val else None
            _c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
        else:
            _c.value = None

    # ── Distribution block ────────────────────────────────────────────────
    _write_distribution_block(ws, issue_keys, settings, eff_header_row)

    # ── Column-label header row ───────────────────────────────────────────
    _write_date_headers(ws, issue_keys, _date_hdr_snap, eff_header_row)

    # ── Merge each column header cell with its padding rows ───────────────
    # Snapshot before unmerge — _unmerge_region pops cells from ws._cells,
    # destroying the values/styles written by _write_date_headers above.
    _hdr_snaps  = {c: _snapshot_style(ws.cell(row=eff_header_row, column=c))
                   for c in range(1, last_col + 1)}
    _hdr_values = {c: ws.cell(row=eff_header_row, column=c).value
                   for c in range(1, last_col + 1)}

    _unmerge_region(ws, eff_header_row, eff_data_start - 1, 1, last_col)
    for _mc in range(1, last_col + 1):
        ws.merge_cells(start_row=eff_header_row, start_column=_mc,
                       end_row=eff_data_start - 1, end_column=_mc)
        _anchor = ws.cell(row=eff_header_row, column=_mc)
        _apply_snapshot(_anchor, _hdr_snaps[_mc])
        _anchor.value = _hdr_values[_mc]

    # ── Clear borders on all interior padding rows (rows 12–15) ──────────
    for _ir in range(eff_header_row + 1, eff_data_start):
        for _ic in range(1, ws.max_column + 1):
            if (_ir, _ic) in ws._cells:
                ws.cell(row=_ir, column=_ic).border = Border()

    # ── Fix yellow border position — full workbook width ─────────────────
    # The template bakes the yellow bottom border into left-column cells at
    # eff_header_row (row 11).  We want it only at eff_data_start-1 (row 15).
    # Strategy: capture the Side style, strip from row 11, stamp onto row 15.
    _yellow_side = None
    for _ic in range(1, ws.max_column + 1):
        _ck = (eff_header_row, _ic)
        if _ck in ws._cells and ws._cells[_ck].has_style:
            _bs = ws._cells[_ck].border.bottom
            if _bs and _bs.style:
                _yellow_side = _bs
                break

    # Strip bottom border from every cell in row 11 across full width
    for _ic in range(1, ws.max_column + 1):
        if (eff_header_row, _ic) in ws._cells:
            _c = ws.cell(row=eff_header_row, column=_ic)
            if _c.has_style and _c.border.bottom.style:
                _b = _c.border
                _c.border = Border(top=_b.top, left=_b.left, right=_b.right)

    # Stamp yellow bottom border onto row 15 across full width
    if _yellow_side is not None:
        for _ic in range(1, ws.max_column + 1):
            _c = ws.cell(row=eff_data_start - 1, column=_ic)
            _b = _c.border if _c.has_style else Border()
            _c.border = Border(top=_b.top, left=_b.left, right=_b.right,
                               bottom=_yellow_side)

    # ── Drawing data rows ─────────────────────────────────────────────────
    last_data_row = _write_data_rows(ws, sheets_data, issue_keys, last_col,
                                     _date_data_snap, eff_data_start, settings=settings)

    # ── Fix merge for rows 1 and 2 to cover the fixed columns ───────────
    # Row 1 merge stops at col K (FIRST_DATE_COL - 1) so the template logo
    # that lives in the right-hand portion of row 1 is not wiped.
    # Row 2 ("DELIVERABLES LIST & ISSUE SHEET") spans the full sheet width.
    _row1_snap = _snapshot_style(ws.cell(row=1, column=1))
    _row1_max_col = ws.max_column  # capture template width before unmerge wipes cells
    _remerge_row(ws, 1, FIRST_DATE_COL - 1)
    _remerge_row(ws, 2, last_col)
    _r1 = ws.cell(row=1, column=1)
    _apply_snapshot(_r1, _row1_snap)
    _r1.value = settings.get('register_title') or project_info.get('project_name', '')
    # Re-apply fill to row 1 date-column area so the logo has a dark background.
    if _row1_snap:
        for _c in range(FIRST_DATE_COL, max(_row1_max_col, 30) + 1):
            ws.cell(row=1, column=_c).fill = _row1_snap['fill']

    # ── Auto-size Drawing Package (col 1), Document Title (col 9), Scale (col 11) ──
    for _cn, _cl in ((1, 'A'), (9, 'I'), (11, 'K')):
        _w = max(
            (len(str(ws.cell(row=_r, column=_cn).value or ''))
             for _r in range(eff_data_start, last_data_row + 1)),
            default=8
        )
        ws.column_dimensions[_cl].width = _w + 3

    # ── Freeze panes ──────────────────────────────────────────────────────
    ws.freeze_panes = ws.cell(row=eff_data_start, column=FIRST_DATE_COL)

    # ── Print area ────────────────────────────────────────────────────────
    ws.print_area = 'A1:{}{}'.format(get_column_letter(last_col), last_data_row)

    # ── Fit everything onto a single page when printing / exporting PDF ───
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0  # unlimited pages tall — scale to fill width only

    # ── Logo: left edge aligned with col I, right edge capped before col L ──
    # Anchor at the left edge of column I (col=8, 0-indexed).
    # Cap logo width so it never crosses into column L (FIRST_DATE_COL).
    # Excel pixel formula: px = round((char_width + 0.71) × 7 + 5)
    ws._images.clear()
    if os.path.exists(_LOGO):
        def _col_px(letter):
            w = ws.column_dimensions[letter].width or 8
            return round((w + 0.71) * 7 + 5)

        _ci_px = _col_px('I')
        _cj_px = _col_px('J')
        _ck_px = _col_px('K')

        _logo_h_px  = 20
        _logo_w_nat = round(753 / 56 * _logo_h_px)          # natural width at aspect ratio
        _logo_w_px  = min(_logo_w_nat, _ci_px + _cj_px + _ck_px - 4)  # cap before col L

        _logo_img        = XLImage(_LOGO)
        _logo_img.height = _logo_h_px
        _logo_img.width  = _logo_w_px
        _marker          = AnchorMarker(col=8, colOff=0,    # col 8 = column I (0-indexed)
                                        row=0, rowOff=101600)
        _size            = XDRPositiveSize2D(cx=_logo_w_px * 9525, cy=_logo_h_px * 9525)
        _logo_img.anchor = OneCellAnchor(_from=_marker, ext=_size)
        ws.add_image(_logo_img)

    # ── Remove any repeat-rows-on-each-page print title set in the template ──
    ws.print_title_rows = None

    wb.template = False
    wb.save(output_path)
    return output_path


# ── Helpers ───────────────────────────────────────────────────────────────────

def _latest_code(sheets_data, date_str, issued_by):
    codes = [
        rev['code']
        for sheet in sheets_data
        for rev in sheet['revisions']
        if rev['date'] == date_str and rev.get('issued_by', '') == issued_by
    ]
    if not codes:
        return ''
    c = [x for x in codes if x.upper().startswith('C')]
    p = [x for x in codes if x.upper().startswith('P')]
    return sorted(c or p or codes)[-1]


def _remerge_row(ws, row, last_col, start_col=1):
    """Remove any existing merge on the row and re-merge from start_col to last_col."""
    _unmerge_region(ws, row, row, start_col, last_col)
    if last_col > start_col:
        ws.merge_cells(start_row=row, start_column=start_col,
                       end_row=row, end_column=last_col)


def _expand_date_columns(ws, template_cols, needed_cols, header_row, data_start):
    """Copy the style of the last template date column to fill in extra columns."""
    ref_col = FIRST_DATE_COL + template_cols - 1
    ref_hdr = ws.cell(row=header_row, column=ref_col)

    for extra in range(template_cols, needed_cols):
        new_col = FIRST_DATE_COL + extra
        # Header row
        _copy_cell_style(ref_hdr, ws.cell(row=header_row, column=new_col))
        ws.column_dimensions[get_column_letter(new_col)].width = \
            ws.column_dimensions[get_column_letter(ref_col)].width or 4.57
        # Distribution rows
        for r in range(_DIST_FIRST_ROW, header_row):
            _copy_cell_style(ws.cell(row=r, column=ref_col),
                             ws.cell(row=r, column=new_col))
        # Data rows
        for r in range(data_start, ws.max_row + 1):
            _copy_cell_style(ws.cell(row=r, column=ref_col),
                             ws.cell(row=r, column=new_col))


def _write_distribution_block(ws, issue_keys, settings, header_row):
    recipients   = settings.get('recipients', [])
    saved_issues = settings.get('issues', {})
    dist_last    = header_row - 1   # last distribution row (inclusive)

    # ── Snapshot A4 rich-text content before unmerge flushes the cell cache ──
    # A4 (the anchor of the left-filler merged block) carries rich-text
    # formatting (bold + coloured runs).  _unmerge_region pops it from
    # ws._cells and the clear loop then sets value=None, destroying it.
    # We snapshot value, style, and original merge range here and restore
    # all three as the very last operation so nothing can overwrite them.
    _a4_cell  = ws.cell(row=_DIST_FIRST_ROW, column=1)
    _a4_value = _a4_cell.value
    _a4_snap  = _snapshot_style(_a4_cell)
    _a4_merge = None
    for _mr in ws.merged_cells.ranges:
        if (_mr.min_row <= _DIST_FIRST_ROW <= _mr.max_row
                and _mr.min_col <= 1 <= _mr.max_col):
            _a4_merge = (_mr.min_row, _mr.min_col, _mr.max_row, _mr.max_col)
            break

    # Capture style refs BEFORE unmerge flushes the cell cache.
    ref_label = ws.cell(row=_DIST_FIRST_ROW, column=10)
    ref_code  = ws.cell(row=_DIST_FIRST_ROW, column=FIRST_DATE_COL)

    # Unmerge and clear all distribution rows.
    # Cols 1-8: also remove template block merges (A4:H6, A7:H10 etc.) that
    #           would leave stray gridlines on unused rows.
    # Cols 9+ : unmerge per-row label merges; preserve template cell borders
    #           so the distribution grid lines remain visible.
    _unmerge_region(ws, _DIST_FIRST_ROW, dist_last, 1, ws.max_column)
    for r in range(_DIST_FIRST_ROW, header_row):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill  = _HEADER_FILL
            cell.value = None
            if c < 9:  # left filler area — no border
                cell.border = Border()

    for i, recipient in enumerate(recipients):
        row    = _DIST_FIRST_ROW + i
        if row > dist_last:
            break   # safety guard — should not happen after row insertion
        r_name = recipient.get('name', '')

        label_cell = ws.cell(row=row, column=9)
        label_cell.value = r_name
        _copy_cell_style(ref_label, label_cell)
        label_cell.fill      = _HEADER_FILL
        label_cell.alignment = Alignment(horizontal='right', vertical='center')
        ws.merge_cells(start_row=row, start_column=9, end_row=row, end_column=11)

        for col_idx, (date_str, issued_by) in enumerate(issue_keys):
            key  = '{}||{}'.format(date_str, issued_by)
            code = saved_issues.get(key, {}).get(r_name, '')
            dc   = FIRST_DATE_COL + col_idx
            code_cell = ws.cell(row=row, column=dc)
            _copy_cell_style(ref_code, code_cell)
            code_cell.fill  = _HEADER_FILL
            code_cell.value = code

    # Ensure every date cell in the last distribution row has a full border
    # (including bottom) so a gridline appears along the bottom edge of the block.
    _ref_border = ref_code.border.copy() if ref_code.has_style else Border()
    _last_dc = max(FIRST_DATE_COL + len(issue_keys) - 1, 35)
    for c in range(FIRST_DATE_COL, _last_dc + 1):
        ws.cell(row=dist_last, column=c).border = _ref_border

    # ── Restore A4 rich-text — must be last so the clear loop cannot clobber it ──
    _a4_out = ws.cell(row=_DIST_FIRST_ROW, column=1)
    _a4_out.value = _a4_value
    _apply_snapshot(_a4_out, _a4_snap)
    if _a4_merge:
        ws.merge_cells(start_row=_a4_merge[0], start_column=_a4_merge[1],
                       end_row=_a4_merge[2], end_column=_a4_merge[3])


def _write_date_headers(ws, issue_keys, date_snap=None, header_row=HEADER_ROW):
    """Write date into the header row date columns, rotated 90°."""
    for col_idx, (date_str, _issued_by) in enumerate(issue_keys):
        col  = FIRST_DATE_COL + col_idx
        cell = ws.cell(row=header_row, column=col)
        _apply_snapshot(cell, date_snap)
        cell.value     = _fmt_header(date_str)
        cell.alignment = Alignment(text_rotation=90, horizontal='center', vertical='bottom')


def _write_data_rows(ws, sheets_data, issue_keys, last_col,
                     date_snap=None, data_start=DATA_ROW_START, settings=None):
    suitability_enabled  = (settings or {}).get('suitability_enabled', False)
    suitability_codes    = (settings or {}).get('suitability_codes', {}) if suitability_enabled else {}
    uncontrolled_enabled = (settings or {}).get('uncontrolled_enabled', False)
    uncontrolled_formats = (settings or {}).get('uncontrolled_formats', {}) if uncontrolled_enabled else {}

    issue_idx = {key: i for i, key in enumerate(issue_keys)}

    # Capture non-date column styles from template data row before unmerge/clear.
    ref_cols = {c: ws.cell(row=data_start, column=c) for c in range(1, FIRST_DATE_COL)}

    # Unmerge and clear all data rows
    _unmerge_region(ws, data_start, ws.max_row, 1, ws.max_column)
    for row in ws.iter_rows(min_row=data_start, max_row=ws.max_row):
        for cell in row:
            cell.value = None

    current_row = data_start
    prev_group  = None

    for sheet in sheets_data:
        group = sheet['sheet_type']

        if prev_group is not None and group != prev_group:
            ws.row_dimensions[current_row].height = _DATA_ROW_HEIGHT
            for _c in range(1, last_col + 1):
                _tgt = ws.cell(row=current_row, column=_c)
                if _c >= FIRST_DATE_COL:
                    _apply_snapshot(_tgt, date_snap)
                else:
                    _src = ref_cols.get(_c, ref_cols.get(min(_c, FIRST_DATE_COL - 1)))
                    if _src is not None:
                        _copy_cell_style(_src, _tgt)
                _tgt.value = None
            current_row += 1

        prev_group = group
        ws.row_dimensions[current_row].height = _DATA_ROW_HEIGHT
        r = current_row

        # Restore styles: left columns from template refs, date columns from snapshot.
        for c in range(1, last_col + 1):
            target = ws.cell(row=r, column=c)
            if c >= FIRST_DATE_COL:
                _apply_snapshot(target, date_snap)
            else:
                src = ref_cols.get(c, ref_cols.get(min(c, FIRST_DATE_COL - 1)))
                if src is not None:
                    _copy_cell_style(src, target)

        values = [
            (1,  sheet['sheet_type'],           'center'),
            (2,  sheet['project'],              'center'),
            (3,  sheet['originator'],           'center'),
            (4,  sheet['functional_breakdown'], 'center'),
            (5,  sheet['spatial_breakdown'],    'center'),
            (6,  sheet['form'],                 'center'),
            (7,  sheet['discipline'],           'center'),
            (8,  sheet['number'],               'center'),
            (9,  sheet['title'],                'left'),
            (10, sheet['size'],                 'center'),
            (11, sheet['scale'],                'center'),
        ]

        for col, val, _ in values:
            cell = ws.cell(row=r, column=col)
            cell.value = val
            if col == 9:  # document title — single line, no wrapping
                cell.alignment = Alignment(
                    horizontal='left', vertical='center', wrap_text=False)

        for rev in sheet['revisions']:
            key = (rev['date'], rev.get('issued_by', ''))
            if key in issue_idx:
                col           = FIRST_DATE_COL + issue_idx[key]
                rev_code      = rev['code']
                issue_key_str = '{}||{}'.format(key[0], key[1])

                if suitability_enabled:
                    suit_code = suitability_codes.get(issue_key_str, {}).get(
                        sheet['sheet_type'], '')
                    cell_val = '{}/{}'.format(rev_code, suit_code) if suit_code else rev_code
                else:
                    cell_val = rev_code

                target_cell       = ws.cell(row=r, column=col)
                target_cell.value = cell_val

                # Red font when this package × issue is marked as uncontrolled
                if uncontrolled_enabled:
                    is_uncontrolled = uncontrolled_formats.get(
                        issue_key_str, {}).get(sheet['sheet_type'], False)
                    if is_uncontrolled:
                        f = target_cell.font
                        target_cell.font = Font(
                            name=f.name, size=f.size, bold=f.bold,
                            italic=f.italic, underline=f.underline,
                            strike=f.strikethrough, color='FFFF0000',
                        )

        current_row += 1

    return current_row - 1  # last row with data
