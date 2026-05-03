# -*- coding: utf-8 -*-
"""
Build the issue register Excel workbook matching the M498 format exactly.

Layout (sheet name: "1.DELIVERABLES LIST")
─────────────────────────────────────────────────────────────────────
Row 1   Project name            (18pt bold, Founders Grotesk Regular)
Row 2   "DELIVERABLES LIST & ISSUE SHEET" + issue date + revision
Rows 3  (empty spacer)
Rows 4-11  Distribution / header block
Row 12  Column headers           (black bg, white 12pt bold)
Row 13+ Drawing data rows       (21.6pt)
─────────────────────────────────────────────────────────────────────

Column layout (header row 12):
  A  Drawing Package   B  Project   C  Originator   D  Functional Breakdown
  E  Spatial Breakdown  F  Form     G  Discipline    H  Number
  I  Document Number   J  Document Title   K  Size   L  Scale
  M… Issue date columns (one per unique date)
"""

from datetime import datetime
from collections import OrderedDict

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_TEXT

# ── Fonts ────────────────────────────────────────────────────────────────────

_FG_REG    = 'Founders Grotesk Regular'
_FG_MED    = 'Founders Grotesk Medium'
_FG_LIGHT  = 'Founders Grotesk Light'
_FALLBACK  = 'Arial'   # used when FG not installed

_BLACK = '000000'
_WHITE = 'FFFFFF'
_DARK2 = '44546A'   # theme dk2 — used for group separator rows
_LIGHT_GREY = 'E7E6E6'  # very light for empty issue cells


def _font(name, size, bold=False, color=_BLACK, italic=False):
    fam = '{}, {}'.format(name, _FALLBACK)
    return Font(name=fam, size=size, bold=bold, color=color, italic=italic)


def _fill(hex_color):
    return PatternFill('solid', fgColor=hex_color)


def _thin_border():
    s = Side(border_style='thin', color=_BLACK)
    return Border(left=s, right=s, top=s, bottom=s)


def _medium_top_border():
    m = Side(border_style='medium', color=_BLACK)
    t = Side(border_style='thin', color=_BLACK)
    return Border(left=t, right=t, top=m, bottom=t)


def _no_border():
    return Border()


# ── Date helpers ─────────────────────────────────────────────────────────────

def _parse_date(date_str):
    """Return datetime or None from common Revit date formats."""
    for fmt in ('%d/%m/%y', '%d/%m/%Y', '%d.%m.%y', '%d.%m.%Y'):
        try:
            return datetime.strptime(date_str, fmt)
        except ValueError:
            pass
    return None


def _format_date_header(date_str):
    """Format issue date as DD/MM/YY for column header."""
    dt = _parse_date(date_str)
    return dt.strftime('%d/%m/%y') if dt else date_str


def _format_date_title(date_str):
    """Format most recent issue date as DD.MM.YYYY for top-right cell."""
    dt = _parse_date(date_str)
    return dt.strftime('%d.%m.%Y') if dt else date_str


# ── Cell writer helpers ───────────────────────────────────────────────────────

def _write(ws, row, col, value,
           font=None, fill=None, align=None, border=None, fmt=None):
    cell = ws.cell(row=row, column=col)
    cell.value = value
    if font   is not None: cell.font      = font
    if fill   is not None: cell.fill      = fill
    if align  is not None: cell.alignment = align
    if border is not None: cell.border    = border
    if fmt    is not None: cell.number_format = fmt
    return cell


def _apply_range(ws, row_start, row_end, col_start, col_end,
                 font=None, fill=None, align=None, border=None):
    """Apply formatting to every cell in a range (does NOT merge)."""
    for r in range(row_start, row_end + 1):
        for c in range(col_start, col_end + 1):
            cell = ws.cell(row=r, column=c)
            if font   is not None: cell.font      = font
            if fill   is not None: cell.fill      = fill
            if align  is not None: cell.alignment = align
            if border is not None: cell.border    = border


# ── Main builder ─────────────────────────────────────────────────────────────

HEADER_ROW   = 12
DATA_ROW_START = 13
FIRST_DATE_COL = 13   # column M (1-based)

# Fixed metadata columns
_COLS = OrderedDict([
    ('drawing_package',      1),   # A
    ('project',              2),   # B
    ('originator',           3),   # C
    ('functional_breakdown', 4),   # D
    ('spatial_breakdown',    5),   # E
    ('form',                 6),   # F
    ('discipline',           7),   # G
    ('number',               8),   # H
    ('doc_number',           9),   # I  (formula)
    ('title',               10),   # J
    ('size',                11),   # K
    ('scale',               12),   # L
])

_HEADER_LABELS = [
    'Drawing Package', 'Project', 'Originator',
    'Functional Breakdown', 'Spatial Breakdown', 'Form', 'Discipline',
    'Number', 'Document Number', 'Document Title', 'Size', 'Scale',
]

# Column widths (characters) matching M498
_COL_WIDTHS = {
    1: 27.3,   # A
    2:  6.6,   # B
    3: 13.0,   # C
    4: 13.0,   # D
    5:  6.6,   # E
    6:  6.6,   # F
    7: 13.0,   # G
    8:  9.1,   # H
    9: 33.9,   # I
   10: 72.0,   # J
   11:  6.6,   # K
   12:  8.6,   # L
   # Date cols (M+): 8.14 each
}
_DATE_COL_WIDTH = 8.14

_ROW_HEIGHTS = {
    1:  25.15,
    2:  50.1,
    **{r: 15.0 for r in range(3, 12)},
    12: 21.6,   # header row
}
_DATA_ROW_HEIGHT = 21.6


def build_register(sheets_data, issue_keys, settings, output_path,
                   project_info):
    """
    Build and save the Excel workbook.

    Parameters
    ----------
    sheets_data   : list of sheet dicts from revit_reader.get_sheets_data()
    issue_keys    : list of (date_str, issued_by) from revit_reader.collect_issue_dates()
    settings      : dict from storage / dialog
    output_path   : full path for the .xlsx file
    project_info  : dict from revit_reader.get_project_info()
    """
    wb = Workbook()
    ws = wb.active
    ws.title = '1.DELIVERABLES LIST'

    n_date_cols = len(issue_keys)
    last_col    = FIRST_DATE_COL + n_date_cols - 1

    # ── Column widths ────────────────────────────────────────────────
    for col_idx, width in _COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    for dc in range(FIRST_DATE_COL, last_col + 1):
        ws.column_dimensions[get_column_letter(dc)].width = _DATE_COL_WIDTH

    # ── Row heights ──────────────────────────────────────────────────
    for row_idx, height in _ROW_HEIGHTS.items():
        ws.row_dimensions[row_idx].height = height

    # ── Row 1: Project name ──────────────────────────────────────────
    _write(ws, 1, 1,
           project_info.get('project_name', ''),
           font=_font(_FG_REG, 18, bold=True),
           align=Alignment(horizontal='left', vertical='center'))
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    ws.row_dimensions[1].height = 25.15

    # ── Row 2: Main title + issue date ───────────────────────────────
    _write(ws, 2, 1,
           'DELIVERABLES LIST & ISSUE SHEET',
           font=_font(_FG_MED, 20),
           align=Alignment(horizontal='left', vertical='center'))
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=10)
    ws.row_dimensions[2].height = 50.1

    _write(ws, 2, 12,
           'Issue date & revision',
           font=_font(_FG_MED, 11),
           align=Alignment(horizontal='right', vertical='center'))

    # Most recent issue date + current revision
    if issue_keys:
        latest_date, latest_by = issue_keys[-1]
        # Find the latest revision code across all sheets on that date
        latest_code = _latest_revision_code(sheets_data, latest_date, latest_by)
        date_rev_text = '{}  |  {}'.format(
            _format_date_title(latest_date), latest_code)
    else:
        date_rev_text = ''

    date_col = 13  # starts at M
    _write(ws, 2, date_col,
           date_rev_text,
           font=_font(_FG_REG, 18, bold=False),
           align=Alignment(horizontal='left', vertical='center'))
    ws.merge_cells(start_row=2, start_column=date_col,
                   end_row=2, end_column=last_col)

    # ── Rows 3: spacer ───────────────────────────────────────────────
    ws.row_dimensions[3].height = 15.0

    # ── Rows 4-11: distribution header block ─────────────────────────
    _write_distribution_block(ws, issue_keys, settings, last_col)

    # ── Row 12: column headers ───────────────────────────────────────
    _write_header_row(ws, issue_keys, last_col)

    # ── Rows 13+: drawing data ───────────────────────────────────────
    _write_data_rows(ws, sheets_data, issue_keys, last_col)

    # ── Freeze panes ─────────────────────────────────────────────────
    ws.freeze_panes = ws.cell(row=DATA_ROW_START, column=FIRST_DATE_COL)

    # ── Page setup for printing ──────────────────────────────────────
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0
    ws.print_area = '{}:{}'.format(
        get_column_letter(1) + '1',
        get_column_letter(last_col) + str(ws.max_row)
    )

    wb.save(output_path)
    return output_path


# ── Sub-sections ─────────────────────────────────────────────────────────────

def _latest_revision_code(sheets_data, date_str, issued_by):
    """Return the latest-sequence revision code issued on the given date."""
    codes = []
    for sheet in sheets_data:
        for rev in sheet['revisions']:
            if rev['date'] == date_str and rev.get('issued_by', '') == issued_by:
                codes.append(rev['code'])
    # Return the 'highest' code (C takes precedence over P)
    if not codes:
        return ''
    c_codes = [c for c in codes if c.upper().startswith('C')]
    p_codes = [c for c in codes if c.upper().startswith('P')]
    pool = c_codes if c_codes else p_codes if p_codes else codes
    return sorted(pool)[-1]


def _write_distribution_block(ws, issue_keys, settings, last_col):
    """
    Write rows 4-11 (distribution matrix header) matching M498 layout.

    Row 4  A4:H6 merged  — KEY text
    Row 4  K4:L4         — "Client" label, M4+ codes
    Row 5  K5:L5         — "Contractor"
    Row 6  I6:L6         — "Project Manager"
    Row 7  L7            — "Structures"
    Row 8  L8            — "MEP"
    Row 9  L9            — "Landscape"
    Row 10 L10           — "CDM PD"
    Row 11 I11:L11       — "Extranet upload"
    """
    key_text = (
        'KEY  Revisions    Pxx Preliminary    Cxx Contractual     '
        'Distribution   E Email   U CDE upload   T Newforma   X Issue slip only    '
        'Format   Controlled in black (pdf, ifc, nwc/d)   Uncontrolled in red (dwg, rvt)   '
        'Status codes (if applicable)    S1 Coordination   S2 Information   '
        'S3 Review & comment   S4 Approval by lead appointing party   '
        'S5 Approval by appointing party   A Approved/Contractual   B Partial sign off'
    )

    key_font  = _font(_FG_MED, 8, bold=True)
    key_align = Alignment(horizontal='left', vertical='top', wrap_text=True)

    _write(ws, 4, 1, key_text, font=key_font, align=key_align)
    ws.merge_cells(start_row=4, start_column=1, end_row=6, end_column=8)

    # Disclaimer text (rows 7-11, cols A-H)
    disclaimer = (
        'The documents listed uncontrolled are issued to enable the recipient to '
        'prepare their own documents/models/drawings for which they are solely '
        'responsible. The documents are based on background information current at '
        'the time of issue. Levitt Bernstein Associates accepts no liability for any '
        'such alterations or additions to or discrepancies arising out of changes to '
        'such background information which occur to that information after it is '
        'issued by Levitt Bernstein Associates.'
    )
    _write(ws, 7, 1, disclaimer,
           font=_font(_FG_LIGHT, 8),
           align=Alignment(horizontal='left', vertical='top', wrap_text=True))
    ws.merge_cells(start_row=7, start_column=1, end_row=11, end_column=8)

    recipients  = settings.get('recipients', [])
    saved_issues = settings.get('issues', {})

    # Map recipient name → row number
    # Rows 4-11 → up to 8 recipients
    recipient_row_map = {
        0: (4,  9, 12),  # Client   label cols 9-12 (I-L)? No, K-L
        1: (5,  9, 12),  # Contractor
        2: (6,  9, 12),  # PM
        3: (7, 12, 12),  # Structures  — label col L only
        4: (8, 12, 12),  # MEP
        5: (9, 12, 12),  # Landscape
        6: (10, 12, 12), # CDM PD
        7: (11,  9, 12), # Extranet upload
    }

    label_font     = _font(_FG_REG, 10)
    label_align_r  = Alignment(horizontal='right', vertical='center')
    label_align_l  = Alignment(horizontal='left',  vertical='center')
    code_font      = _font(_FG_MED, 10, bold=False)
    code_align     = Alignment(horizontal='center', vertical='center')
    tb             = _thin_border()

    for i, recipient in enumerate(recipients[:8]):
        r_name = recipient.get('name', '')
        if i not in recipient_row_map:
            break
        excel_row, label_start, label_end = recipient_row_map[i]

        # Merge and write label
        if label_start < label_end:
            ws.merge_cells(start_row=excel_row, start_column=label_start,
                           end_row=excel_row, end_column=label_end)
        _write(ws, excel_row, label_start, r_name,
               font=label_font, align=label_align_r, border=tb)

        # Apply thin border to merged label range
        for mc in range(label_start, label_end + 1):
            ws.cell(row=excel_row, column=mc).border = tb

        # Distribution codes per issue column
        for col_idx, (date_str, issued_by) in enumerate(issue_keys):
            key = '{}||{}'.format(date_str, issued_by)
            code = saved_issues.get(key, {}).get(r_name, '')
            dc = FIRST_DATE_COL + col_idx
            _write(ws, excel_row, dc, code,
                   font=code_font, align=code_align, border=tb)

    # Apply thin border to all distribution cells that have no value yet
    for dist_row in range(4, 12):
        for dc in range(FIRST_DATE_COL, last_col + 1):
            cell = ws.cell(row=dist_row, column=dc)
            if cell.border is None or cell.border == Border():
                cell.border = tb


def _write_header_row(ws, issue_keys, last_col):
    """Row 12: black background, white bold 12pt column headers."""
    hdr_font  = _font(_FG_MED, 12, bold=True, color=_WHITE)
    hdr_fill  = _fill(_BLACK)
    hdr_align = Alignment(horizontal='center', vertical='center',
                          wrap_text=True)
    tb = _thin_border()

    ws.row_dimensions[HEADER_ROW].height = 21.6

    for i, label in enumerate(_HEADER_LABELS):
        col = i + 1
        _write(ws, HEADER_ROW, col, label,
               font=hdr_font, fill=hdr_fill, align=hdr_align, border=tb)

    for col_idx, (date_str, issued_by) in enumerate(issue_keys):
        col = FIRST_DATE_COL + col_idx
        header_text = _format_date_header(date_str)
        if issued_by:
            header_text += '\n' + issued_by
        _write(ws, HEADER_ROW, col, header_text,
               font=hdr_font, fill=hdr_fill, align=hdr_align, border=tb)


def _write_data_rows(ws, sheets_data, issue_keys, last_col):
    """Write all drawing rows, with blank separator rows between groups."""
    issue_key_index = {key: idx for idx, key in enumerate(issue_keys)}

    data_font   = _font(_FG_REG, 11)
    data_align  = Alignment(horizontal='left', vertical='center')
    center_align = Alignment(horizontal='center', vertical='center')
    rev_align   = Alignment(horizontal='center', vertical='center')
    tb          = _thin_border()

    current_row = DATA_ROW_START
    prev_group  = None

    for sheet in sheets_data:
        group = sheet['sheet_type']

        # Blank separator between groups
        if prev_group is not None and group != prev_group:
            ws.row_dimensions[current_row].height = _DATA_ROW_HEIGHT
            current_row += 1

        prev_group = group

        ws.row_dimensions[current_row].height = _DATA_ROW_HEIGHT

        # Build the doc number formula
        # =IF(B13="", "", CONCATENATE(B13,"-",C13,"-",D13,"-",E13,"-",F13,"-",G13,"-",H13))
        r = current_row
        formula = (
            '=IF(B{r}="","",CONCATENATE('
            'B{r},"-",C{r},"-",D{r},"-",E{r},"-",F{r},"-",G{r},"-",H{r}))'
        ).format(r=r)

        values = [
            (1,  sheet['sheet_type'],           center_align),
            (2,  sheet['project'],              center_align),
            (3,  sheet['originator'],           center_align),
            (4,  sheet['functional_breakdown'], center_align),
            (5,  sheet['spatial_breakdown'],    center_align),
            (6,  sheet['form'],                 center_align),
            (7,  sheet['discipline'],           center_align),
            (8,  sheet['number'],               center_align),
            (9,  formula,                       data_align),
            (10, sheet['title'],                data_align),
            (11, sheet['size'],                 center_align),
            (12, sheet['scale'],                center_align),
        ]

        for col, val, al in values:
            _write(ws, r, col, val,
                   font=data_font, align=al, border=tb)

        # Revision codes in date columns
        for rev in sheet['revisions']:
            key = (rev['date'], rev.get('issued_by', ''))
            if key in issue_key_index:
                col = FIRST_DATE_COL + issue_key_index[key]
                _write(ws, r, col, rev['code'],
                       font=data_font, align=rev_align, border=tb)

        # Empty cells with border for date columns that have no revision
        for col_idx in range(len(issue_keys)):
            col = FIRST_DATE_COL + col_idx
            cell = ws.cell(row=r, column=col)
            if cell.value is None:
                cell.border = tb

        current_row += 1
